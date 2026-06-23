# Copyright (c) 2026, Frappe Technologies Pvt. Ltd. and contributors
# For license information, please see license.txt

import io
import re

import frappe
from frappe import _
from frappe.model.document import Document
from frappe.utils.file_manager import get_file

# Seed data for the master doctypes. The Dutch labels/aliases are the exact strings
# (case/whitespace-insensitive) that appear in the Manuren reference sheet and drive
# the column/row matching during import.
VARIANT_SEED = [
	(1, "Standard", "Standaard"),
	(2, "4WD E-Diff", "4WD/E-sper"),
	(3, "R CUPRA", "R/CUPRA"),
	(4, "RS VRS CUPRA 280-90", "RS/VRS/CUPRA 280-90"),
	(5, "6-8 Cylinder", "6/8 cillinder"),
]

JOB_SEED = [
	# job_name, dutch_aliases, diagnosis_keywords, project_part_fields, item_match_tokens, use_oe_crossref, use_oe_pn_search, use_transmission_search
	("Oil Change", ["Oliewissel"], ["oil change", "olie", "oliewissel"], [], ["oil", "olie", "oliewissel"], 0, 0, 0),
	("Mechatronics", ["Mechatronic vervangen", "Mechatronic vervangen / TCU", "Mechatronic"], ["mechatronic", "tcu"], ["mechatronic"], ["mec", "mechatronic", "tcu"], 0, 1, 0),
	("Clutch", ["Koppeling vervangen"], ["clutch", "koppeling"], ["clutch"], ["clu", "kop", "koppeling", "clutch"], 1, 0, 0),
	("Flywheel", ["Vliegwiel vervangen"], ["flywheel", "vliegwiel"], ["flywheel"], ["fly", "vli", "vliegwiel", "flywheel"], 1, 0, 0),
	("Gearbox", ["Versnellingsbak vervangen"], ["gearbox", "transmission", "versnellingsbak"], ["dsg_gearbox"], ["bak", "gear", "versnellingsbak", "gearbox"], 0, 0, 1),
	("Clutch + Flywheel", ["Koppeling + Vliegwiel vervangen"], ["clutch and flywheel", "koppeling + vliegwiel"], ["clutch", "flywheel"], ["clu", "kop", "koppeling", "fly", "vli", "vliegwiel"], 1, 0, 0),
	("Ride-height Sensor", ["Rijstand sensor vervangen"], ["ride height sensor", "rijstand", "level sensor"], [], ["rijstand", "niveau", "level", "sensor"], 0, 0, 0),
]

DSG_CODE_RE = re.compile(r"^(dq|dl)\d{3}$")


def _norm(value):
	"""Lower-case, collapse internal whitespace, strip ends."""
	if value is None:
		return ""
	return re.sub(r"\s+", " ", str(value).strip().lower())


def parse_value(raw):
	"""Parse a Manuren cell into (applicable, hours).

	"1,5 uur" -> (True, 1.5), "7 uur " -> (True, 7.0), "NVT" -> (False, None),
	blank -> (False, None).
	"""
	if raw is None:
		return (False, None)
	s = str(raw).strip()
	if not s or "nvt" in s.lower():
		return (False, None)
	s = s.lower().replace("uur", " ").replace(",", ".")
	m = re.search(r"\d+(?:\.\d+)?", s)
	if not m:
		return (False, None)
	return (True, float(m.group()))


def ensure_masters():
	"""Create the Vehicle Variant and Labour Job master records if missing."""
	for no, name, dutch in VARIANT_SEED:
		if not frappe.db.exists("Vehicle Variant", name):
			frappe.get_doc(
				{
					"doctype": "Vehicle Variant",
					"variant_no": no,
					"variant_name": name,
					"dutch_label": dutch,
				}
			).insert(ignore_permissions=True)

	for name, aliases, keywords, part_fields, item_tokens, use_crossref, use_pn_search, use_trans in JOB_SEED:
		if not frappe.db.exists("Labour Job", name):
			frappe.get_doc(
				{
					"doctype": "Labour Job",
					"job_name": name,
					"dutch_aliases": "\n".join(aliases),
					"match_keywords": "\n".join(keywords),
					"project_part_fields": ", ".join(part_fields),
					"item_match_keywords": ", ".join(item_tokens),
					"use_oe_crossref": use_crossref,
					"use_oe_pn_search": use_pn_search,
					"use_transmission_search": use_trans,
				}
			).insert(ignore_permissions=True)
		else:
			# Backfill mappings on jobs seeded before these fields existed.
			if part_fields and not frappe.db.get_value("Labour Job", name, "project_part_fields"):
				frappe.db.set_value("Labour Job", name, "project_part_fields", ", ".join(part_fields))
			if item_tokens and not frappe.db.get_value("Labour Job", name, "item_match_keywords"):
				frappe.db.set_value("Labour Job", name, "item_match_keywords", ", ".join(item_tokens))
			if use_crossref and not frappe.db.get_value("Labour Job", name, "use_oe_crossref"):
				frappe.db.set_value("Labour Job", name, "use_oe_crossref", use_crossref)
			if use_pn_search and not frappe.db.get_value("Labour Job", name, "use_oe_pn_search"):
				frappe.db.set_value("Labour Job", name, "use_oe_pn_search", use_pn_search)
			if use_trans and not frappe.db.get_value("Labour Job", name, "use_transmission_search"):
				frappe.db.set_value("Labour Job", name, "use_transmission_search", use_trans)


def _build_lookups():
	"""Build normalised lookup dicts from the current master records."""
	variant_by_dutch = {}
	for v in frappe.get_all("Vehicle Variant", fields=["variant_name", "dutch_label"]):
		if v.dutch_label:
			variant_by_dutch[_norm(v.dutch_label)] = v.variant_name

	job_by_alias = {}
	for j in frappe.get_all("Labour Job", fields=["job_name", "dutch_aliases"]):
		for alias in re.split(r"[\n,]", j.dutch_aliases or ""):
			alias = _norm(alias)
			if alias:
				job_by_alias[alias] = j.job_name

	return variant_by_dutch, job_by_alias


def _load_rows(file_url):
	"""Return the worksheet rows (with data) as a list of value-lists."""
	_, content = get_file(file_url)
	if isinstance(content, str):
		content = content.encode("utf-8")

	import openpyxl

	wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True, read_only=True)
	ws = wb["Sheet1"] if "Sheet1" in wb.sheetnames else wb.worksheets[0]
	return [list(row) for row in ws.iter_rows(values_only=True)]


def parse_sheet(rows, variant_by_dutch, job_by_alias):
	"""Walk the sheet by content anchors and return (entries, warnings).

	Handles the two stacked halves and a section's DSG header / variant header
	appearing anywhere, so it does not depend on fixed row numbers.
	"""
	entries = []
	warnings = []
	col_map = {}  # column index -> variant_name (for the current half)
	current_dsg = None

	for idx, values in enumerate(rows, start=1):
		# Variant header row: two or more cells match known Dutch variant labels.
		matched_cols = {}
		for ci, val in enumerate(values):
			nv = _norm(val)
			if nv and nv in variant_by_dutch:
				matched_cols[ci] = variant_by_dutch[nv]
		if len(matched_cols) >= 2:
			col_map = matched_cols
			continue

		a = _norm(values[0]) if values else ""
		if not a:
			continue

		if DSG_CODE_RE.match(a):
			current_dsg = str(values[0]).strip().upper()
			continue

		if a in job_by_alias:
			job = job_by_alias[a]
			if not current_dsg or not col_map:
				warnings.append(f"Row {idx}: job '{values[0]}' found before a DSG/variant header — skipped.")
				continue
			for ci, variant in col_map.items():
				raw = values[ci] if ci < len(values) else None
				applicable, hours = parse_value(raw)
				entries.append(
					{
						"dsg_code": current_dsg,
						"job": job,
						"vehicle_variant": variant,
						"applicable": 1 if applicable else 0,
						"hours": hours or 0,
					}
				)

	return entries, warnings


class ManurenImport(Document):
	@frappe.whitelist()
	def run_import(self):
		if not self.attachment:
			frappe.throw(_("Attach a Manuren .xlsx file first."))
		if not self.year:
			frappe.throw(_("Set the Year first."))

		log = []
		try:
			ensure_masters()
			variant_by_dutch, job_by_alias = _build_lookups()
			rows = _load_rows(self.attachment)
			entries, warnings = parse_sheet(rows, variant_by_dutch, job_by_alias)
			log.extend(warnings)

			created = updated = 0
			seen_dsg = set()
			for e in entries:
				# Auto-create missing DSG Code masters.
				if e["dsg_code"] not in seen_dsg:
					seen_dsg.add(e["dsg_code"])
					if not frappe.db.exists("DSG Code", e["dsg_code"]):
						frappe.get_doc({"doctype": "DSG Code", "code": e["dsg_code"]}).insert(
							ignore_permissions=True
						)
						log.append(f"Created DSG Code '{e['dsg_code']}'.")

				existing = frappe.db.exists(
					"Standard Labour Hours",
					{
						"year": self.year,
						"dsg_code": e["dsg_code"],
						"job": e["job"],
						"vehicle_variant": e["vehicle_variant"],
					},
				)
				if existing:
					doc = frappe.get_doc("Standard Labour Hours", existing)
					updated += 1
				else:
					doc = frappe.new_doc("Standard Labour Hours")
					doc.year = self.year
					doc.dsg_code = e["dsg_code"]
					doc.job = e["job"]
					doc.vehicle_variant = e["vehicle_variant"]
					created += 1

				doc.applicable = e["applicable"]
				doc.hours = e["hours"]
				doc.source_import = self.name
				doc.save(ignore_permissions=True)

			self.rows_created = created + updated
			self.status = "Imported"
			log.insert(0, f"Imported {created + updated} rows ({created} new, {updated} updated).")
			self.log = "\n".join(log)
			self.save(ignore_permissions=True)
			return {"created": created, "updated": updated, "warnings": warnings}

		except Exception:
			self.status = "Failed"
			self.log = "\n".join(log + [frappe.get_traceback()])
			self.save(ignore_permissions=True)
			raise
